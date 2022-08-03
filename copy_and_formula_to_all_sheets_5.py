#ANURAJ PILANKU
#Copy styles of cells  and implement it on  another cell to all sheets
#transfering formula  to all sheets
import openpyxl
import sys
from openpyxl.formula.translate import Translator
from copy import copy


wb = openpyxl.load_workbook(sys.argv[1])
def changes(sheetnum):
    newsheet = wb.worksheets[sheetnum]
    row = newsheet.max_row - 1

    def name(columnnum):
        cell = newsheet.cell(row=row, column=columnnum)
        new_cell = newsheet.cell(row=row + 1, column=columnnum)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            # new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
            # new_cell.value = copy(cell.value)

    name(1)
    name(2)
    name(3)
    name(4)
    name(5)
    name(6)
    name(7)

    def set(columnnum):
        cell = newsheet.cell(row=row, column=columnnum)
        new_cell = newsheet.cell(row=row + 1, column=columnnum)
        if cell.has_style:
            new_cell.value = copy(cell.value)
    set(1)
    set(4)
    set(5)
    set(6)
    set(7)

    # transfering formula from a cell to another cell
    newsheet["A" + str(row + 1)] = Translator("=IPMPDF!A"+str(row),origin="A" + str(row)).translate_formula("A" + str(row + 1))

    newsheet["D" + str(row + 1)] = Translator("=B" + str(row) + "-C" + str(row),
                                              origin="D" + str(row)).translate_formula("D" + str(row + 1))
    newsheet["E" + str(row + 1)] = Translator("=D" + str(row) + "/B" + str(row),
                                              origin="E" + str(row)).translate_formula("E" + str(row + 1))
    newsheet["F" + str(row + 1)] = Translator("=C" + str(row) + "-C" + str(row - 1),
                                              origin="F" + str(row)).translate_formula("F" + str(row + 1))
for i in range(2,len(wb.worksheets)-1):
    changes(i)
wb.save(sys.argv[1])
print("success")

#python kwargs.py C:\Users\2040664\anuraj\IPM_FSM\fsm.xlsx
