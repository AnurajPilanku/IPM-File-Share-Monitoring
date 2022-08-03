# ANURAJ PILANKU
# IPM FILE SHARE MONITORING USECASE
# COPYING DATA FROM NOTEPAD TO EXISTING EXCEL SHEET
# note:if data is already present in excel , this code will note overwrite it ,the data will be placed after the existing data

#importing libraries
import _csv#_csv
import sys

import numpy as np
import openpyxl
import numpy

#commandline arguments
notefile=sys.argv[1]
excelfile=sys.argv[2]

#opening excel workbook
workkbook=openpyxl.load_workbook(excelfile)#.Workbook()
sheet=workkbook.worksheets[14]
transferlist=list()
with open(notefile, 'r') as file:
    records = _csv.reader(file)#records = _csv.reader(file)
    next(records)
    for record in records:
        #removing all blank spaces in a string
        firstline=record[0]
        sec=firstline.split("\t")
        #remove all empty items in a list
        filtered=list(filter(None,sec))
        #if there is space between the first two words
        if any(c.isalpha() for c in filtered[1]):
            filtered[0]=filtered[0]+filtered[1]#filtered[0]+" "+filtered[1]
            filtered.pop(filtered.index(filtered[1]))
        firstpart=filtered[:1]
        secondpart=list(map(float,(filtered[1:])))#changing datatpye of a list from string to float
        roundvalues = np.round_(secondpart, decimals=2)
        firstpart.extend(roundvalues)#combine two lists of different datatypes
        if firstpart[0] not in ["CCase"]:
            transferlist.append(firstpart)
#copying data to excel sheet
for j in range(1,4):
    for i in range(2,len(transferlist)+2):
        sheet.cell(row=i,column=j).value=transferlist[i-2][j-1]
workkbook.save(excelfile)
print("success")


#copying data to excel sheet
# for i in transferlist:
#     sheet.append(i)
# workkbook.save(excelfile)
# print("sucsess")




#python pyautogui.py C:\Users\2040664\anuraj\IPM_FSM\DiskMonitoring.txt C:\Users\2040664\anuraj\IPM_FSM\test.xlsx



