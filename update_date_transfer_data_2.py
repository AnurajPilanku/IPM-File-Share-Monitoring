#ANURAJ PILANKU
#Update date and update space details from reference sheet
import datetime
import sys
import openpyxl
from datetime import date, timedelta

#commandline arguments
excelfile=sys.argv[1]

#opening excel workbook
workbook=openpyxl.load_workbook(excelfile)#.Workbook()
ipmpdf=workbook.worksheets[1]
reference=workbook.worksheets[14]
row=ipmpdf.max_row
#copy-pasting data from previous row to next in each column and changing date #7-Jan-2021
#change Date
changed_date=str(ipmpdf.cell(row=row,column=1).value + timedelta(days=7))
format_change=datetime.datetime.strptime(changed_date,"%Y-%m-%d %H:%M:%S")
requiredformat=format_change.strftime("%d-%m-%Y")
finalised=datetime.datetime.strptime(requiredformat,"%d-%m-%Y")
ipmpdf.cell(row=row+1,column=1).value=requiredformat#ipmpdf.cell(row=row,column=1).value + timedelta(days=7)
#changing the format of a cell
ipmpdf.cell(row=row+1,column=1).number_format="dd-mm-yy"
#total space
ipmpdf.cell(row=row+1,column=2).value=reference.cell(row=2,column=2).value
#available space
ipmpdf.cell(row=row+1,column=3).value=reference.cell(row=2,column=3).value
workbook.save(excelfile)
print("success")

#python csv.py C:\Users\2040664\anuraj\IPM_FSM\fsm.xlsx
