#ANURAJ PILANKU
#update space details from reference sheet to all other sheets
import datetime
import sys
import openpyxl

#commandline arguments
excelfile=sys.argv[1]

#opening excel workbook
workbook=openpyxl.load_workbook(excelfile)#.Workbook()
reference=workbook.worksheets[14]
def sheets(sheetnum):
    ipmpdf = workbook.worksheets[sheetnum]
    row = ipmpdf.max_row
    ipmpdf.cell(row=row + 1, column=2).value = reference.cell(row=sheetnum+1, column=2).value
    # available space
    ipmpdf.cell(row=row + 1, column=3).value = reference.cell(row=sheetnum+1, column=3).value
for i in range(2,len(workbook.worksheets)-1):
   sheets(i)

workbook.save(excelfile)
print("success")

#python csv.py C:\Users\2040664\anuraj\IPM_FSM\fsm.xlsx
